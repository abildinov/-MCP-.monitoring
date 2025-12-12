"""
Excel report generator
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any
from pathlib import Path
import tempfile
import os
from loguru import logger

from .data_collector import collect_metrics_for_period


async def generate_excel_report(
    period: str,
    output_path: str = None,
    prometheus_url: str = "http://147.45.157.2:9090",
    loki_url: str = "http://147.45.157.2:3100"
) -> str:
    """
    Генерирует Excel отчёт за указанный период
    
    Args:
        period: Период ("24h", "7d", "30d")
        output_path: Путь для сохранения файла
        prometheus_url: URL Prometheus
        loki_url: URL Loki
        
    Returns:
        Путь к созданному файлу
    """
    logger.info(f"Генерация Excel отчёта за период: {period}")
    
    # Собираем данные
    data = await collect_metrics_for_period(period, prometheus_url, loki_url)
    
    # Создаём путь к файлу если не указан
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Используем системную временную директорию (работает и на Windows и на Linux)
        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, f"report_{period}_{timestamp}.xlsx")
    
    # Создаём Excel workbook
    wb = openpyxl.Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Создаём листы отчёта
    create_summary_sheet(wb, data)
    create_cpu_sheet(wb, data)
    create_memory_sheet(wb, data)
    create_disk_sheet(wb, data)
    create_network_sheet(wb, data)
    create_alerts_sheet(wb, data)
    create_errors_sheet(wb, data)
    create_processes_sheet(wb, data)
    
    # Сохраняем файл
    wb.save(output_path)
    logger.info(f"Отчёт сохранён: {output_path}")
    
    return output_path


def create_summary_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Summary с краткой сводкой"""
    ws = wb.create_sheet("Summary", 0)
    
    # Заголовок
    ws['A1'] = 'ОТЧЁТ О СИСТЕМЕ МОНИТОРИНГА'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Информация о периоде
    ws['A3'] = 'Период:'
    ws['B3'] = data['period']
    ws['A4'] = 'Начало:'
    ws['B4'] = data['start_time']
    ws['A5'] = 'Конец:'
    ws['B5'] = data['end_time']
    
    # Заголовок секции метрик
    ws['A7'] = 'ОСНОВНЫЕ МЕТРИКИ'
    ws['A7'].font = Font(size=14, bold=True)
    
    # Заголовки таблицы
    headers = ['Метрика', 'Минимум', 'Среднее', 'Медиана', 'P95', 'Максимум', 'Текущее', 'Тренд', 'Статус']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # CPU
    cpu = data['cpu']
    row = 9
    ws.cell(row, 1, 'CPU Usage (%)')
    ws.cell(row, 2, round(cpu.get('min', 0), 2))
    ws.cell(row, 3, round(cpu.get('avg', 0), 2))
    ws.cell(row, 4, round(cpu.get('median', 0), 2))
    ws.cell(row, 5, round(cpu.get('p95', 0), 2))
    ws.cell(row, 6, round(cpu.get('max', 0), 2))
    ws.cell(row, 7, round(cpu.get('current', 0), 2))
    ws.cell(row, 8, cpu.get('trend', 'N/A'))
    ws.cell(row, 9, get_status(cpu.get('current', 0), 80))
    apply_status_color(ws.cell(row, 9), cpu.get('current', 0), 80)
    
    # Memory
    memory = data['memory']
    row = 10
    ws.cell(row, 1, 'Memory Usage (%)')
    ws.cell(row, 2, round(memory.get('min', 0), 2))
    ws.cell(row, 3, round(memory.get('avg', 0), 2))
    ws.cell(row, 4, round(memory.get('median', 0), 2))
    ws.cell(row, 5, round(memory.get('p95', 0), 2))
    ws.cell(row, 6, round(memory.get('max', 0), 2))
    ws.cell(row, 7, round(memory.get('current', 0), 2))
    ws.cell(row, 8, memory.get('trend', 'N/A'))
    ws.cell(row, 9, get_status(memory.get('current', 0), 85))
    apply_status_color(ws.cell(row, 9), memory.get('current', 0), 85)
    
    # Диски
    ws['A12'] = 'ДИСКИ'
    ws['A12'].font = Font(size=12, bold=True)
    
    row = 13
    if data['disk']['disks']:
        for disk in data['disk']['disks']:
            ws.cell(row, 1, f"{disk['mountpoint']}")
            ws.cell(row, 2, f"{disk['percent']:.1f}%")
            ws.cell(row, 3, get_status(disk['percent'], 90))
            apply_status_color(ws.cell(row, 3), disk['percent'], 90)
            row += 1
    
    # Алерты
    row += 1
    ws.cell(row, 1, 'АЛЕРТЫ')
    ws.cell(row, 1).font = Font(size=12, bold=True)
    row += 1
    
    alerts = data['alerts']
    if alerts:
        firing_alerts = [a for a in alerts if a.get('state') in ['firing_now', 'firing']]
        historical_alerts = [a for a in alerts if a.get('state') == 'fired_in_period']
        
        ws.cell(row, 1, f'Активных алертов: {len(firing_alerts)}')
        if firing_alerts:
            ws.cell(row, 1).font = Font(color='FF0000', bold=True)
        row += 1
        ws.cell(row, 1, f'Алертов за период: {len(historical_alerts)}')
        ws.cell(row, 1).font = Font(color='FFA500' if historical_alerts else '00AA00')
    else:
        ws.cell(row, 1, 'Алертов нет')
        ws.cell(row, 1).font = Font(color='00AA00')
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 20  # Метрика
    ws.column_dimensions['B'].width = 10  # Минимум
    ws.column_dimensions['C'].width = 10  # Среднее
    ws.column_dimensions['D'].width = 10  # Медиана
    ws.column_dimensions['E'].width = 10  # P95
    ws.column_dimensions['F'].width = 10  # Максимум
    ws.column_dimensions['G'].width = 10  # Текущее
    ws.column_dimensions['H'].width = 15  # Тренд
    ws.column_dimensions['I'].width = 12  # Статус


def create_cpu_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист CPU"""
    ws = wb.create_sheet("CPU")
    
    ws['A1'] = 'CPU USAGE ANALYSIS'
    ws['A1'].font = Font(size=16, bold=True)
    
    cpu = data['cpu']
    
    # Заголовок статистики
    ws['A3'] = 'СТАТИСТИКА ЗА ПЕРИОД'
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:D3')
    
    # Заголовки таблицы статистики
    stats_headers = ['Метрика', 'Значение', 'Статус', 'Тренд']
    for col, header in enumerate(stats_headers, start=1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Данные статистики
    row = 5
    stats = [
        ('Текущее', cpu.get('current', 0), 80),
        ('Среднее', cpu.get('avg', 0), 80),
        ('Медиана', cpu.get('median', 0), 80),
        ('Минимум', cpu.get('min', 0), 80),
        ('Максимум', cpu.get('max', 0), 80),
        ('95 процентиль', cpu.get('p95', 0), 80),
    ]
    
    for metric, value, threshold in stats:
        ws.cell(row, 1, metric)
        ws.cell(row, 2, f"{value:.2f}%")
        ws.cell(row, 3, get_status(value, threshold))
        apply_status_color(ws.cell(row, 3), value, threshold)
        
        # Тренд только для текущего
        if metric == 'Текущее':
            ws.cell(row, 4, cpu.get('trend', 'N/A'))
        
        row += 1
    
    # Количество замеров
    row += 1
    ws.cell(row, 1, 'Замеров:')
    ws.cell(row, 2, cpu.get('samples', 0))
    ws.cell(row, 1).font = Font(italic=True)
    
    # Временной ряд
    row += 2
    ws.cell(row, 1, 'ДИНАМИКА ПО ВРЕМЕНИ (последние 48 точек)')
    ws.cell(row, 1).font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    # Заголовки
    ws.cell(row, 1, 'Время')
    ws.cell(row, 2, 'CPU %')
    ws.cell(row, 3, 'Статус')
    for col in range(1, 4):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Данные - показываем последние 48 точек (4 часа при 5-минутном интервале)
    values = cpu.get('values', [])
    timestamps = cpu.get('timestamps', [])
    
    # Берём последние 48 точек
    display_values = values[-48:] if len(values) > 48 else values
    display_timestamps = timestamps[-48:] if len(timestamps) > 48 else timestamps
    
    row += 1
    for timestamp, value in zip(display_timestamps, display_values):
        ws.cell(row, 1, timestamp.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row, 2, round(value, 2))
        ws.cell(row, 3, get_status(value, 80))
        apply_status_color(ws.cell(row, 3), value, 80)
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_memory_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Memory"""
    ws = wb.create_sheet("Memory")
    
    ws['A1'] = 'MEMORY USAGE ANALYSIS'
    ws['A1'].font = Font(size=16, bold=True)
    
    memory = data['memory']
    
    # Общая информация
    ws['A3'] = 'Общая память:'
    ws['B3'] = f"{memory.get('total_gb', 0):.2f} GB"
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'].font = Font(size=11)
    
    # Статистика
    ws['A5'] = 'СТАТИСТИКА ЗА ПЕРИОД'
    ws['A5'].font = Font(size=12, bold=True)
    ws.merge_cells('A5:D5')
    
    # Заголовки
    stats_headers = ['Метрика', 'Значение', 'Статус', 'Тренд']
    for col, header in enumerate(stats_headers, start=1):
        cell = ws.cell(6, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Данные статистики
    row = 7
    stats = [
        ('Текущее', memory.get('current', 0), 85),
        ('Среднее', memory.get('avg', 0), 85),
        ('Медиана', memory.get('median', 0), 85),
        ('Минимум', memory.get('min', 0), 85),
        ('Максимум', memory.get('max', 0), 85),
        ('95 процентиль', memory.get('p95', 0), 85),
    ]
    
    for metric, value, threshold in stats:
        ws.cell(row, 1, metric)
        ws.cell(row, 2, f"{value:.2f}%")
        ws.cell(row, 3, get_status(value, threshold))
        apply_status_color(ws.cell(row, 3), value, threshold)
        
        # Тренд
        if metric == 'Текущее':
            ws.cell(row, 4, memory.get('trend', 'N/A'))
        
        row += 1
    
    # Количество замеров
    row += 1
    ws.cell(row, 1, 'Замеров:')
    ws.cell(row, 2, memory.get('samples', 0))
    ws.cell(row, 1).font = Font(italic=True)
    
    # Временной ряд
    row += 2
    ws.cell(row, 1, 'ДИНАМИКА ПО ВРЕМЕНИ (последние 48 точек)')
    ws.cell(row, 1).font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws.cell(row, 1, 'Время')
    ws.cell(row, 2, 'Memory %')
    ws.cell(row, 3, 'Статус')
    for col in range(1, 4):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Данные
    values = memory.get('values', [])
    timestamps = memory.get('timestamps', [])
    
    display_values = values[-48:] if len(values) > 48 else values
    display_timestamps = timestamps[-48:] if len(timestamps) > 48 else timestamps
    
    row += 1
    for timestamp, value in zip(display_timestamps, display_values):
        ws.cell(row, 1, timestamp.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row, 2, round(value, 2))
        ws.cell(row, 3, get_status(value, 85))
        apply_status_color(ws.cell(row, 3), value, 85)
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_disk_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Disk"""
    ws = wb.create_sheet("Disk")
    
    ws['A1'] = 'DISK USAGE'
    ws['A1'].font = Font(size=14, bold=True)
    
    # Заголовки
    headers = ['Device', 'Mountpoint', 'Used %', 'Used GB', 'Total GB', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True)
    
    # Данные дисков
    row = 4
    for disk in data['disk'].get('disks', []):
        ws.cell(row, 1, disk.get('device', 'unknown'))
        ws.cell(row, 2, disk.get('mountpoint', '/'))
        ws.cell(row, 3, round(disk.get('percent', 0), 2))
        ws.cell(row, 4, round(disk.get('used_gb', 0), 2))
        ws.cell(row, 5, round(disk.get('total_gb', 0), 2))
        ws.cell(row, 6, get_status(disk.get('percent', 0), 90))
        apply_status_color(ws.cell(row, 6), disk.get('percent', 0), 90)
        row += 1
    
    # IO статистика
    row += 2
    ws.cell(row, 1, 'IO Statistics (avg)')
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 1, 'Read:')
    ws.cell(row, 2, f"{data['disk'].get('io_read_avg_mb', 0):.2f} MB/s")
    row += 1
    ws.cell(row, 1, 'Write:')
    ws.cell(row, 2, f"{data['disk'].get('io_write_avg_mb', 0):.2f} MB/s")
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_network_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Network"""
    ws = wb.create_sheet("Network")
    
    ws['A1'] = 'NETWORK STATUS'
    ws['A1'].font = Font(size=14, bold=True)
    
    network = data['network']
    
    ws['A3'] = 'Общая информация'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Status:'
    ws['B4'] = network.get('status', 'unknown')
    ws['A5'] = 'Interfaces:'
    ws['B5'] = network.get('interfaces', 0)
    
    ws['A7'] = 'Трафик (average)'
    ws['A7'].font = Font(bold=True)
    ws['A8'] = 'RX:'
    ws['B8'] = f"{network.get('rx_avg_mb', 0):.2f} MB/s"
    ws['A9'] = 'TX:'
    ws['B9'] = f"{network.get('tx_avg_mb', 0):.2f} MB/s"
    
    ws['A11'] = 'Ошибки (average)'
    ws['A11'].font = Font(bold=True)
    ws['A12'] = 'Errors/sec:'
    ws['B12'] = f"{network.get('errors_avg', 0):.2f}"
    
    ws['A14'] = 'Соединения'
    ws['A14'].font = Font(bold=True)
    connections = network.get('connections', {})
    ws['A15'] = 'TCP established:'
    ws['B15'] = connections.get('tcp_established', 0)
    ws['A16'] = 'UDP datagrams:'
    ws['B16'] = connections.get('udp_datagrams', 0)
    ws['A17'] = 'Total:'
    ws['B17'] = connections.get('total', 0)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


def create_alerts_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Alerts"""
    ws = wb.create_sheet("Alerts")
    
    ws['A1'] = 'ALERTS HISTORY'
    ws['A1'].font = Font(size=14, bold=True)
    
    # Заголовки
    headers = ['Alert Name', 'Severity', 'Status', 'Firing Count', 'First Fired', 'Last Fired']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Данные алертов
    row = 4
    alerts = data.get('alerts', [])
    
    if alerts:
        # Сортируем: сначала активные (firing_now), потом по количеству срабатываний
        sorted_alerts = sorted(
            alerts,
            key=lambda x: (x.get('state') != 'firing_now', -x.get('firing_count', 0))
        )
        
        for alert in sorted_alerts:
            ws.cell(row, 1, alert.get('name', 'Unknown'))
            ws.cell(row, 2, alert.get('severity', 'unknown'))
            
            # Статус
            state = alert.get('state', 'unknown')
            state_text = {
                'firing_now': '🔴 ACTIVE NOW',
                'fired_in_period': '⚠️ Fired in period',
                'unknown': 'Unknown'
            }.get(state, state)
            ws.cell(row, 3, state_text)
            
            # Количество срабатываний
            firing_count = alert.get('firing_count', 0)
            ws.cell(row, 4, firing_count)
            
            # Время первого и последнего срабатывания
            ws.cell(row, 5, alert.get('first_fired', 'N/A'))
            ws.cell(row, 6, alert.get('last_fired', 'N/A'))
            
            # Подсветка активных алертов
            if state == 'firing_now':
                for col in range(1, 7):
                    ws.cell(row, col).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    ws.cell(row, col).font = Font(bold=True)
            elif firing_count > 10:
                # Если было много срабатываний - оранжевый
                for col in range(1, 7):
                    ws.cell(row, col).fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            
            row += 1
    
    if row == 4:
        ws.cell(4, 1, 'Нет алертов за указанный период')
        ws.cell(4, 1).font = Font(color='00AA00')
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 25  # Alert Name
    ws.column_dimensions['B'].width = 12  # Severity
    ws.column_dimensions['C'].width = 18  # Status
    ws.column_dimensions['D'].width = 15  # Firing Count
    ws.column_dimensions['E'].width = 20  # First Fired
    ws.column_dimensions['F'].width = 20  # Last Fired


def create_errors_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Errors"""
    ws = wb.create_sheet("Errors")
    
    ws['A1'] = 'ERRORS FROM LOGS'
    ws['A1'].font = Font(size=14, bold=True)
    
    # Заголовки
    headers = ['Timestamp', 'Container', 'Message']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True)
    
    # Данные ошибок
    row = 4
    for error in data.get('errors', []):
        ws.cell(row, 1, error.get('timestamp', ''))
        ws.cell(row, 2, error.get('container', 'unknown'))
        ws.cell(row, 3, error.get('message', ''))
        row += 1
    
    if row == 4:
        ws.cell(4, 1, 'Нет ошибок')
        ws.cell(4, 1).font = Font(color='00AA00')
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60


def create_processes_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]):
    """Создаёт лист Processes"""
    ws = wb.create_sheet("Top Processes")
    
    ws['A1'] = 'TOP PROCESSES BY CPU'
    ws['A1'].font = Font(size=14, bold=True)
    
    # Заголовки
    headers = ['Rank', 'Process Name', 'CPU Usage %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True)
    
    # Данные процессов
    row = 4
    for process in data.get('processes', []):
        ws.cell(row, 1, process.get('rank', row - 3))
        ws.cell(row, 2, process.get('name', 'unknown'))
        ws.cell(row, 3, round(process.get('cpu_usage', 0), 2))
        row += 1
    
    if row == 4:
        ws.cell(4, 1, 'Нет данных о процессах')
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15


def get_status(value: float, threshold: float) -> str:
    """Возвращает статус на основе значения и порога"""
    if value >= threshold:
        return 'HIGH'
    elif value >= threshold * 0.8:
        return 'WARNING'
    else:
        return 'NORMAL'


def apply_status_color(cell, value: float, threshold: float):
    """Применяет цвет ячейки в зависимости от статуса"""
    if value >= threshold:
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    elif value >= threshold * 0.8:
        cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        cell.font = Font(bold=True)
    else:
        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        cell.font = Font(bold=True)

